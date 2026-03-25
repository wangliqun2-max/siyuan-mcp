"""打印 Checklist 模板前150行的原始内容"""
import openpyxl, os

xlsx_path = os.path.join('backend', '..', 'Checklist模板.xlsx')
wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active

print(f"共 {ws.max_row} 行, {ws.max_column} 列")
print()
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    # Print all non-empty rows
    vals = [str(v) if v is not None else '' for v in row[:5]]
    if any(v.strip() for v in vals):
        print(f"Row {i:3d}: {' | '.join(vals)}")
    if i > 150:
        break
