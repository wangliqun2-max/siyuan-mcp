"""
excel_handler.py - Read Excel checklist template and write extracted parameters

Template structure (Checklist模板.xlsx):
- Column C (3): Item number (1.1, 1.2, 2.1 etc.)
- Column D (4): Parameter name (Chinese/English)
- Column L (12): Value to be filled in (投标值) — left side params
- Column U (21): Item number for right-side params (some rows have 2 params per row)
- Column V (22): Parameter name for right-side params
- Column AD (30): Value to be filled in for right-side params
- Section headers are rows where Col C contains text like "1.运行环境", "2.基本参数" (no col D)
"""
import openpyxl
from openpyxl.cell import MergedCell
from openpyxl.styles import PatternFill, Font, Alignment
import io

# Column indices (1-based)
COL_LEFT_NUM = 3    # C: item number (left)
COL_LEFT_NAME = 4   # D: param name (left)
COL_LEFT_VALUE = 12  # L: value to fill (left)
COL_RIGHT_NUM = 21   # U: item number (right)
COL_RIGHT_NAME = 22  # V: param name (right)
COL_RIGHT_VALUE = 30  # AD: value to fill (right)

# Color fills for status
FILL_FOUND = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
FILL_NOT_FOUND = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")


def read_template_params(template_path: str) -> dict:
    """
    Read the Excel checklist template and extract all parameter names with their locations.

    Returns:
        {
            "sheet_name": str,
            "params": [
                {
                    "name": str,        # Clean parameter label
                    "row": int,         # Row number in sheet
                    "value_col": int,   # Column index to write extracted value
                    "side": "left"|"right"
                }
            ]
        }
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # ── Step 1: Scan section header rows (col C has title, col D is empty) ───
    # These are rows like row 13: colC="1.运行环境", colD=EMPTY
    section_headers = []   # [{"row": int, "title": str, "section_num": int}]
    DATA_START_ROW = 13
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        c_val = ws.cell(row=row_idx, column=COL_LEFT_NUM).value    # col C
        d_val = ws.cell(row=row_idx, column=COL_LEFT_NAME).value   # col D
        if c_val and (d_val is None or not str(d_val).strip()):
            title = str(c_val).strip()
            # Skip noise rows like "部件名称"
            if any(ch.isdigit() for ch in title):
                sec_num = _section_from_item_num(title)
                section_headers.append({"row": row_idx, "title": title, "section_num": sec_num})

    def _get_section_title(row_idx: int) -> str:
        """Return the most recent section header title that precedes row_idx."""
        title = ""
        for sh in section_headers:
            if sh["row"] <= row_idx:
                title = sh["title"]
            else:
                break
        return title

    # ── Step 2: Read params ──────────────────────────────────────────────────
    params = []

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        # --- LEFT side: item in col C, name in col D ---
        left_num_raw = ws.cell(row=row_idx, column=COL_LEFT_NUM).value
        left_name = ws.cell(row=row_idx, column=COL_LEFT_NAME).value
        left_num = str(left_num_raw).strip() if left_num_raw else ""

        # Skip rows with no item number (unnumbered extra rows in template)
        if left_name and str(left_name).strip() and left_num:
            clean_name = str(left_name).strip().replace('\n', ' ')
            params.append({
                "name": clean_name,
                "item_num": left_num,
                "section_num": _section_from_item_num(left_num),
                "section_title": _get_section_title(row_idx),
                "row": row_idx,
                "value_col": COL_LEFT_VALUE,
                "side": "left"
            })

        # --- RIGHT side: item in col U, name in col V ---
        right_num_raw = ws.cell(row=row_idx, column=COL_RIGHT_NUM).value
        right_name = ws.cell(row=row_idx, column=COL_RIGHT_NAME).value
        right_num = str(right_num_raw).strip() if right_num_raw else ""

        if right_name and str(right_name).strip() and right_num:
            clean_right = str(right_name).strip().replace('\n', ' ')
            params.append({
                "name": clean_right,
                "item_num": right_num,
                "section_num": _section_from_item_num(right_num),
                "section_title": _get_section_title(row_idx),
                "row": row_idx,
                "value_col": COL_RIGHT_VALUE,
                "side": "right"
            })

    # Sort by: (section number, left-before-right, row)
    # This groups section-1 left (1.1-1.5) → section-1 right (1.6-1.10) → section-2 left → ...
    params.sort(key=lambda p: (
        _section_from_item_num(p.get("item_num", "")),
        0 if p.get("side") == "left" else 1,
        p.get("row", 0),
    ))

    # Deduplicate param names: if two params share a name, append [item_num] to disambiguate.
    # This prevents dict key collisions when building extracted_ordered in app.py.
    from collections import Counter
    name_count = Counter(p["name"] for p in params)
    name_seen: dict[str, int] = {}
    for p in params:
        name = p["name"]
        if name_count[name] > 1:
            name_seen[name] = name_seen.get(name, 0) + 1
            num = p.get("item_num", "").strip()
            p["name"] = f"{name} [{num}]" if num else f"{name} ({name_seen[name]})"

    # Build sections list for frontend grouping
    sections = []
    current_section = None
    for p in params:
        if p.get("section_title") != (current_section["title"] if current_section else None):
            current_section = {"title": p.get("section_title", ""), "params": []}
            sections.append(current_section)
        current_section["params"].append(p["name"])

    return {
        "sheet_name": ws.title,
        "params": params,
        "sections": sections,
    }



def write_results_to_excel(template_path: str, extracted_params: dict) -> bytes:
    """
    Write LLM-extracted parameters into a copy of the Excel template.

    Args:
        template_path: Path to the Excel template
        extracted_params: Dict from LLM extractor:
            {"param_name": {"value": "...", "unit": "...", "source_text": "...", "found": bool}}

    Returns:
        bytes: Excel file content for download
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    template_info = read_template_params(template_path)

    matched_count = 0
    for param_info in template_info["params"]:
        param_name = param_info["name"]
        row = param_info["row"]
        value_col = param_info["value_col"]

        matched_data = _find_match(param_name, extracted_params)

        if matched_data and matched_data.get("found"):
            value_str = str(matched_data.get("value", ""))
            unit = matched_data.get("unit", "")
            if unit and unit.lower() not in ["n/a", "none", "null", "-", ""]:
                value_str = f"{value_str} {unit}".strip()
            _safe_write_cell(ws, row, value_col, value=value_str, fill=FILL_FOUND)
            matched_count += 1

        elif matched_data and not matched_data.get("found"):
            _safe_write_cell(ws, row, value_col, value="—", fill=FILL_NOT_FOUND)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _section_from_item_num(item_num: str) -> int:
    """Return the integer section prefix of an item number (e.g. '2.4' → 2, '15.1' → 15)."""
    if not item_num or not item_num.strip():
        return 999
    try:
        return int(item_num.strip().split(".")[0].strip())
    except (ValueError, IndexError):
        return 999


def _parse_item_num(item_num: str) -> tuple:
    """
    Convert an item number string like "1.2.3" or "15.1" to a numeric tuple
    for sorting, so that 1.2 < 1.10 < 2.1 < 15.1.
    Items with no number sort to the end.
    """
    if not item_num or not item_num.strip():
        return (999999,)
    try:
        return tuple(int(x) for x in item_num.strip().split(".") if x.strip().isdigit())
    except ValueError:
        return (999999,)


def _safe_write_cell(ws, row: int, col: int, value=None, fill=None) -> None:
    """
    Write value/fill to a worksheet cell, safely handling merged cells.

    When *cell* is a MergedCell (not the top-left of its merged region),
    openpyxl's .value attribute is read-only.  This function locates the
    top-left (master) cell of the merge range and writes there instead.
    """
    cell = ws.cell(row=row, column=col)

    if isinstance(cell, MergedCell):
        # Find which merged range this cell belongs to and use its top-left cell
        for merge_range in ws.merged_cells.ranges:
            if (
                merge_range.min_row <= row <= merge_range.max_row
                and merge_range.min_col <= col <= merge_range.max_col
            ):
                cell = ws.cell(row=merge_range.min_row, column=merge_range.min_col)
                break
        else:
            return  # Can't resolve — skip silently

    if value is not None:
        cell.value = value
    if fill is not None:
        try:
            cell.fill = fill
        except Exception:
            pass  # Some edge-case cells don't accept fill; skip silently


def _find_match(param_name: str, extracted_params: dict) -> dict | None:
    """Find a matching parameter using exact, then fuzzy name matching."""
    if not extracted_params:
        return None

    # Exact match
    if param_name in extracted_params:
        return extracted_params[param_name]

    # Case-insensitive exact match
    param_lower = param_name.lower()
    for key, val in extracted_params.items():
        if key.lower() == param_lower:
            return val

    # Substring match (either direction)
    for key, val in extracted_params.items():
        key_lower = key.lower()
        if param_lower in key_lower or key_lower in param_lower:
            return val

    # Token overlap match (for long bilingual params like "Full Wave Lightning Impulse 雷电全波冲击")
    param_tokens = set(param_lower.replace('(', ' ').replace(')', ' ').split())
    for key, val in extracted_params.items():
        key_tokens = set(key.lower().replace('(', ' ').replace(')', ' ').split())
        overlap = param_tokens & key_tokens
        if len(overlap) >= 2 and len(overlap) / max(len(param_tokens), 1) > 0.4:
            return val

    return None
